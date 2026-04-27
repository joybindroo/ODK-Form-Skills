import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import json

def extract_template_metadata(template_path):
    """
    Extracts the column structures (headers) from the template to provide 
    the AI agent with a precise schema for form creation.
    """
    if not os.path.exists(template_path):
        return None
    
    try:
        wb = load_workbook(template_path, data_only=True)
        metadata = {}
        
        target_sheets = ['survey', 'choices', 'settings', 'entities']
        for sheet in target_sheets:
            if sheet in wb.sheetnames:
                ws = wb[sheet]
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]]
                metadata[sheet] = headers
        return metadata
    except Exception as e:
        print(f"Error extracting metadata: {e}")
        return None

def generate_xlsform(output_path, survey_data, choices_data, settings_data, template_path=None):
    """
    Generates a professional ODK XLSForm .xlsx file.
    
    Args:
        output_path (str): Path to save the .xlsx file.
        survey_data (list): List of rows for the survey sheet.
        choices_data (dict or list): 
            - If dict: { 'list_name': [ { 'name': '...', 'label': '...', 'filter_col': 'val' }, ... ] }
            - If list: Standard list of rows including headers.
        settings_data (dict): Dictionary of settings.
        template_path (str, optional): Path to a template.
    """
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    # 1. Survey Sheet
    ws_survey = wb.create_sheet('survey', 0)
    if isinstance(survey_data, list) and len(survey_data) > 0:
        # If the first row is not a header (e.g. just data), we add standard ODK headers
        if not any(isinstance(cell, str) and cell.lower() == 'type' for cell in survey_data[0]):
            ws_survey.append(['type', 'name', 'label', 'appearance', 'required', 'constraint', 'relevance', 'choice_filter'])
        
        for row in survey_data:
            ws_survey.append(row)

    # 2. Choices Sheet
    ws_choices = wb.create_sheet('choices', 1)
    
    if isinstance(choices_data, dict):
        # Advanced Mode: Handle dictionary with cascading filter support
        # Determine all unique filter columns across all choice lists
        all_filter_cols = set()
        for list_name, options in choices_data.items():
            for opt in options:
                for key in opt.keys():
                    if key not in ['name', 'label']:
                        all_filter_cols.add(key)
        
        filter_cols = sorted(list(all_filter_cols))
        headers = ['list_name', 'name', 'label'] + filter_cols
        ws_choices.append(headers)
        
        for list_name, options in choices_data.items():
            for opt in options:
                row = [list_name, opt.get('name', ''), opt.get('label', '')]
                for col in filter_cols:
                    row.append(opt.get(col, ''))
                ws_choices.append(row)
    else:
        # Legacy Mode: Standard list of rows
        for row in choices_data:
            ws_choices.append(row)

    # 3. Settings Sheet
    ws_settings = wb.create_sheet('settings', 2)
    if isinstance(settings_data, dict):
        ws_settings.append(list(settings_data.keys()))
        ws_settings.append(list(settings_data.values()))
    elif isinstance(settings_data, list):
        for row in settings_data:
            ws_settings.append(row)

    wb.save(output_path)

if __name__ == "__main__":
    # Example usage for testing
    settings = {'form_title': 'Enhanced Test Form', 'form_id': 'enhanced_v1', 'version': '1'}
    survey = [
        ['select_one state', 'state', 'Select State', '', 'yes', '', '', ''],
        ['select_one city', 'city', 'Select City', '', 'yes', '', '', 'state = ${state}']
    ]
    choices = {
        'state': [
            {'name': 'S1', 'label': 'State 1'},
            {'name': 'S2', 'label': 'State 2'}
        ],
        'city': [
            {'name': 'C1', 'label': 'City 1', 'state': 'S1'},
            {'name': 'C2', 'label': 'City 2', 'state': 'S1'},
            {'name': 'C3', 'label': 'City 3', 'state': 'S2'}
        ]
    }
    
    generate_xlsform('enhanced_test.xlsx', survey, choices, settings)
    print("Enhanced XLSForm generated successfully.")
